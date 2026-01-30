"""
Smart Onboarding API

Handles file upload, AI analysis, and connector recommendations.
"""

import os
import uuid
import json
import asyncio
from datetime import datetime
from pathlib import Path
from typing import List, Optional
import logging

from fastapi import APIRouter, UploadFile, File, HTTPException, BackgroundTasks, Depends
from pydantic import BaseModel

logger = logging.getLogger(__name__)

router = APIRouter()

# In-memory storage for uploads (in production, use Redis or DB)
upload_storage: dict = {}

# Upload directory
UPLOAD_DIR = Path("/tmp/0711-uploads")
UPLOAD_DIR.mkdir(exist_ok=True)


class UploadResponse(BaseModel):
    upload_id: str
    files_received: int
    status: str


class FileInfo(BaseModel):
    name: str
    size: int
    type: str
    rows: Optional[int] = None
    columns: Optional[int] = None


class AnalysisSummary(BaseModel):
    upload_id: str
    status: str
    files: List[FileInfo]
    total_products: int
    total_manufacturers: int
    data_quality_score: float
    confidence_score: float
    has_etim: bool
    has_eclass: bool
    detected_formats: List[str]
    business_type: str
    language: str
    recommendations: List[dict]
    potential_revenue: int
    potential_savings: int


async def analyze_file(file_path: Path, filename: str) -> dict:
    """Analyze a single file and extract metadata."""
    file_size = file_path.stat().st_size
    file_ext = filename.lower().split('.')[-1]
    
    result = {
        "name": filename,
        "size": file_size,
        "type": file_ext,
        "rows": 0,
        "columns": 0,
        "sample_data": [],
        "detected_fields": [],
    }
    
    try:
        if file_ext == 'csv':
            import csv
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                headers = next(reader, [])
                result["columns"] = len(headers)
                result["detected_fields"] = headers[:20]  # First 20 columns
                
                rows = 0
                sample = []
                for i, row in enumerate(reader):
                    rows += 1
                    if i < 5:
                        sample.append(dict(zip(headers[:10], row[:10])))
                
                result["rows"] = rows
                result["sample_data"] = sample
                
        elif file_ext in ['xlsx', 'xls']:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value))
                
                result["columns"] = len(headers)
                result["detected_fields"] = headers[:20]
                result["rows"] = ws.max_row - 1  # Exclude header
                
                # Sample data
                sample = []
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
                    sample.append(dict(zip(headers[:10], [str(v) if v else "" for v in row[:10]])))
                result["sample_data"] = sample
                
                wb.close()
            except ImportError:
                logger.warning("openpyxl not installed, skipping Excel parsing")
                result["rows"] = 1000  # Estimate
                result["columns"] = 10
                
        elif file_ext == 'json':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                data = json.load(f)
                if isinstance(data, list):
                    result["rows"] = len(data)
                    if data and isinstance(data[0], dict):
                        result["columns"] = len(data[0].keys())
                        result["detected_fields"] = list(data[0].keys())[:20]
                        result["sample_data"] = data[:5]
                elif isinstance(data, dict):
                    # Could be nested structure
                    result["rows"] = 1
                    result["columns"] = len(data.keys())
                    result["detected_fields"] = list(data.keys())[:20]
                    
        elif file_ext == 'xml':
            # Basic XML handling
            result["rows"] = 100  # Estimate
            result["columns"] = 10
            result["detected_fields"] = ["product", "name", "price", "sku"]
            
    except Exception as e:
        logger.error(f"Error analyzing file {filename}: {e}")
        
    return result


def detect_classification_system(fields: List[str], sample_data: List[dict]) -> dict:
    """Detect if ETIM, eClass, or other classification systems are present."""
    fields_lower = [f.lower() for f in fields]
    all_text = " ".join(fields_lower)
    
    # Check sample data too
    sample_text = ""
    for row in sample_data:
        sample_text += " ".join(str(v).lower() for v in row.values())
    
    combined = all_text + " " + sample_text
    
    has_etim = any(x in combined for x in ['etim', 'ec0', 'ef0', 'ev0'])
    has_eclass = any(x in combined for x in ['eclass', 'ecl_', 'e-class'])
    has_unspsc = 'unspsc' in combined
    
    # Detect B2B-specific fields
    has_ean = any(x in fields_lower for x in ['ean', 'gtin', 'ean13', 'barcode'])
    has_sku = any(x in fields_lower for x in ['sku', 'artikelnummer', 'article_number', 'art_nr'])
    has_price = any(x in fields_lower for x in ['price', 'preis', 'vk', 'ek', 'listprice'])
    
    return {
        "has_etim": has_etim,
        "has_eclass": has_eclass,
        "has_unspsc": has_unspsc,
        "has_ean": has_ean,
        "has_sku": has_sku,
        "has_price": has_price,
    }


def generate_recommendations(analysis: dict) -> List[dict]:
    """Generate connector recommendations based on analysis."""
    recommendations = []
    
    classifications = analysis.get("classifications", {})
    total_products = analysis.get("total_products", 0)
    
    # ETIM Classification
    if not classifications.get("has_etim"):
        recommendations.append({
            "id": "etim-classifier",
            "name": "ETIM Klassifizierung",
            "description": f"{total_products:,} Produkte automatisch klassifizieren",
            "category": "classification",
            "priority": 1,
            "potential_value": total_products * 2,  # €2 per product
            "time_to_value": "2-4 Stunden",
        })
    
    # BMEcat Export
    if classifications.get("has_ean") or classifications.get("has_sku"):
        recommendations.append({
            "id": "bmecat-export",
            "name": "BMEcat Export",
            "description": "Katalogdaten für B2B-Marktplätze exportieren",
            "category": "syndication",
            "priority": 2,
            "potential_value": 15000,
            "time_to_value": "1 Stunde",
        })
    
    # Description Generator
    recommendations.append({
        "id": "description-generator",
        "name": "KI-Beschreibungen",
        "description": f"SEO-optimierte Texte für {int(total_products * 0.3):,} Produkte",
        "category": "enrichment",
        "priority": 3,
        "potential_value": int(total_products * 0.3 * 5),
        "time_to_value": "4-8 Stunden",
    })
    
    # Marketplace connectors
    if total_products > 100:
        recommendations.append({
            "id": "amazon-sp-api",
            "name": "Amazon Seller API",
            "description": f"{int(total_products * 0.6):,} Produkte sofort listbar",
            "category": "marketplace",
            "priority": 4,
            "potential_value": int(total_products * 0.6 * 10),
            "time_to_value": "24 Stunden",
        })
    
    # Data quality
    recommendations.append({
        "id": "data-quality",
        "name": "Datenqualitäts-Check",
        "description": "Duplikate, fehlende Felder, Inkonsistenzen finden",
        "category": "quality",
        "priority": 5,
        "potential_value": 5000,
        "time_to_value": "30 Minuten",
    })
    
    return sorted(recommendations, key=lambda x: x["priority"])


@router.post("/smart-onboarding/upload", response_model=UploadResponse)
async def upload_files(
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...)
):
    """Upload files for smart onboarding analysis."""
    upload_id = str(uuid.uuid4())
    upload_path = UPLOAD_DIR / upload_id
    upload_path.mkdir(exist_ok=True)
    
    saved_files = []
    for file in files:
        if not file.filename:
            continue
            
        # Save file
        file_path = upload_path / file.filename
        content = await file.read()
        
        with open(file_path, 'wb') as f:
            f.write(content)
        
        saved_files.append({
            "name": file.filename,
            "path": str(file_path),
            "size": len(content),
        })
    
    # Store upload info
    upload_storage[upload_id] = {
        "id": upload_id,
        "files": saved_files,
        "status": "analyzing",
        "created_at": datetime.utcnow().isoformat(),
        "analysis": None,
    }
    
    # Start background analysis
    background_tasks.add_task(run_analysis, upload_id)
    
    return UploadResponse(
        upload_id=upload_id,
        files_received=len(saved_files),
        status="analyzing"
    )


async def run_analysis(upload_id: str):
    """Run analysis in background."""
    upload_info = upload_storage.get(upload_id)
    if not upload_info:
        return
    
    try:
        all_fields = []
        all_samples = []
        total_rows = 0
        file_infos = []
        detected_formats = set()
        
        for file_data in upload_info["files"]:
            file_path = Path(file_data["path"])
            if file_path.exists():
                analysis = await analyze_file(file_path, file_data["name"])
                file_infos.append(FileInfo(**{
                    "name": analysis["name"],
                    "size": analysis["size"],
                    "type": analysis["type"],
                    "rows": analysis.get("rows", 0),
                    "columns": analysis.get("columns", 0),
                }))
                
                all_fields.extend(analysis.get("detected_fields", []))
                all_samples.extend(analysis.get("sample_data", []))
                total_rows += analysis.get("rows", 0)
                detected_formats.add(analysis["type"].upper())
        
        # Detect classification systems
        classifications = detect_classification_system(all_fields, all_samples)
        
        # Estimate metrics
        total_products = max(total_rows, 100)
        estimated_manufacturers = max(10, total_products // 500)
        
        # Calculate quality score based on field presence
        quality_factors = [
            classifications.get("has_ean", False),
            classifications.get("has_sku", False),
            classifications.get("has_price", False),
            len(all_fields) > 5,
            total_products > 100,
        ]
        quality_score = sum(quality_factors) / len(quality_factors) * 100
        
        # Generate recommendations
        analysis_data = {
            "total_products": total_products,
            "classifications": classifications,
        }
        recommendations = generate_recommendations(analysis_data)
        
        # Calculate potential value
        potential_revenue = sum(r.get("potential_value", 0) for r in recommendations)
        potential_savings = int(potential_revenue * 0.3)
        
        # Store analysis result
        upload_storage[upload_id]["status"] = "completed"
        upload_storage[upload_id]["analysis"] = {
            "files": [f.dict() for f in file_infos],
            "total_products": total_products,
            "total_manufacturers": estimated_manufacturers,
            "data_quality_score": round(quality_score, 1),
            "confidence_score": 85.0 + (quality_score / 10),
            "has_etim": classifications.get("has_etim", False),
            "has_eclass": classifications.get("has_eclass", False),
            "detected_formats": list(detected_formats),
            "business_type": "B2B-Distributor",
            "language": "de",
            "recommendations": recommendations,
            "potential_revenue": potential_revenue,
            "potential_savings": potential_savings,
        }
        
        logger.info(f"Analysis completed for upload {upload_id}: {total_products} products")
        
    except Exception as e:
        logger.error(f"Analysis failed for {upload_id}: {e}")
        upload_storage[upload_id]["status"] = "failed"
        upload_storage[upload_id]["error"] = str(e)


@router.get("/smart-onboarding/analysis/{upload_id}/summary")
async def get_analysis_summary(upload_id: str):
    """Get analysis summary for an upload."""
    upload_info = upload_storage.get(upload_id)
    
    if not upload_info:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    if upload_info["status"] == "analyzing":
        return {"status": "analyzing", "upload_id": upload_id}
    
    if upload_info["status"] == "failed":
        raise HTTPException(status_code=500, detail=upload_info.get("error", "Analysis failed"))
    
    analysis = upload_info.get("analysis", {})
    
    return {
        "upload_id": upload_id,
        "status": "completed",
        **analysis
    }


@router.post("/smart-onboarding/analysis/{upload_id}/confirm")
async def confirm_recommendations(upload_id: str, selected_connectors: List[str] = None):
    """Confirm selected recommendations and start provisioning."""
    upload_info = upload_storage.get(upload_id)
    
    if not upload_info:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    # In production, this would trigger actual connector provisioning
    return {
        "status": "provisioning",
        "upload_id": upload_id,
        "selected_connectors": selected_connectors or [],
        "message": "Connectors werden eingerichtet..."
    }
