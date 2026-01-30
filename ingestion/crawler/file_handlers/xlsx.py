"""
XLSX Handler - Extract data from Excel spreadsheets
"""

import asyncio
from pathlib import Path
from typing import Optional
import logging

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from .base import BaseHandler

logger = logging.getLogger(__name__)


class XLSXHandler(BaseHandler):
    """
    Extract data from Excel files (.xlsx, .xlsm).

    Converts spreadsheet data to structured text format.
    Handles multiple sheets and preserves basic structure.
    """

    @classmethod
    def supported_extensions(cls) -> set[str]:
        return {'.xlsx', '.xlsm'}

    def __init__(self, max_rows: int = 10000, max_cols: int = 100):
        """
        Args:
            max_rows: Maximum rows to process per sheet
            max_cols: Maximum columns to process per sheet
        """
        self.max_rows = max_rows
        self.max_cols = max_cols

    async def extract(self, path: Path) -> Optional[str]:
        """Extract data from Excel file"""
        if load_workbook is None:
            logger.error("openpyxl not installed, cannot extract .xlsx files")
            return None

        return await asyncio.get_event_loop().run_in_executor(
            None, self._extract_sync, path
        )

    def _extract_sync(self, path: Path) -> Optional[str]:
        """Synchronous extraction"""
        try:
            # Load workbook (data_only=True to get calculated values)
            wb = load_workbook(str(path), data_only=True, read_only=True)

            text_parts = []

            # Process each sheet
            for sheet_name in wb.sheetnames:
                try:
                    sheet = wb[sheet_name]

                    # Extract sheet data
                    sheet_text = self._extract_sheet(sheet, sheet_name)

                    if sheet_text:
                        text_parts.append(sheet_text)

                except Exception as e:
                    logger.warning(f"Failed to extract sheet '{sheet_name}' from {path}: {e}")
                    continue

            wb.close()

            # Combine all sheets
            full_text = "\n\n".join(text_parts)

            if not full_text.strip():
                logger.info(f"No data in Excel file: {path}")
                return None

            return full_text

        except Exception as e:
            logger.error(f"Excel extraction failed for {path}: {e}")
            return None

    def _extract_sheet(self, sheet, sheet_name: str) -> str:
        """Extract data from a single sheet"""
        rows = []

        # Add sheet header
        rows.append(f"=== Sheet: {sheet_name} ===\n")

        # Get dimensions
        max_row = min(sheet.max_row, self.max_rows) if sheet.max_row else 0
        max_col = min(sheet.max_column, self.max_cols) if sheet.max_column else 0

        if max_row == 0 or max_col == 0:
            return ""

        # Extract rows
        row_count = 0
        for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            # Get cell values
            values = []
            for cell in row:
                value = cell.value

                # Convert value to string
                if value is None:
                    values.append("")
                elif isinstance(value, (int, float)):
                    values.append(str(value))
                else:
                    values.append(str(value).strip())

            # Skip completely empty rows
            if not any(v for v in values):
                continue

            # Format as TSV (tab-separated)
            rows.append("\t".join(values))
            row_count += 1

        if row_count == 0:
            return ""

        return "\n".join(rows)


class XLSHandler(BaseHandler):
    """
    Handler for legacy Excel files (.xls).

    Requires xlrd library or conversion to .xlsx.
    """

    @classmethod
    def supported_extensions(cls) -> set[str]:
        return {'.xls'}

    async def extract(self, path: Path) -> Optional[str]:
        """Extract data from legacy Excel file"""
        logger.warning(f"Legacy .xls format detected: {path}")

        # Try xlrd first (if available)
        try:
            import xlrd
            return await self._extract_with_xlrd(path)
        except ImportError:
            logger.info("xlrd not available, attempting conversion to .xlsx")

        # Fall back to conversion
        converted_path = await self._convert_to_xlsx(path)

        if converted_path and converted_path.exists():
            handler = XLSXHandler()
            text = await handler.extract(converted_path)

            # Cleanup
            converted_path.unlink(missing_ok=True)

            return text

        return None

    async def _extract_with_xlrd(self, path: Path) -> Optional[str]:
        """Extract using xlrd library"""
        import xlrd

        return await asyncio.get_event_loop().run_in_executor(
            None, self._extract_xlrd_sync, path
        )

    def _extract_xlrd_sync(self, path: Path) -> Optional[str]:
        """Synchronous xlrd extraction"""
        try:
            import xlrd

            wb = xlrd.open_workbook(str(path))
            text_parts = []

            for sheet in wb.sheets():
                rows = [f"=== Sheet: {sheet.name} ===\n"]

                for row_idx in range(sheet.nrows):
                    values = []
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        values.append(str(cell.value) if cell.value else "")

                    if any(v for v in values):
                        rows.append("\t".join(values))

                if len(rows) > 1:
                    text_parts.append("\n".join(rows))

            return "\n\n".join(text_parts) if text_parts else None

        except Exception as e:
            logger.error(f"xlrd extraction failed for {path}: {e}")
            return None

    async def _convert_to_xlsx(self, path: Path) -> Optional[Path]:
        """Convert .xls to .xlsx using LibreOffice"""
        try:
            import subprocess
            import tempfile

            temp_dir = Path(tempfile.mkdtemp())
            output_path = temp_dir / f"{path.stem}.xlsx"

            cmd = [
                "libreoffice",
                "--headless",
                "--convert-to", "xlsx",
                "--outdir", str(temp_dir),
                str(path)
            ]

            process = await asyncio.create_subprocess_exec(
                *cmd,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE
            )

            await asyncio.wait_for(process.communicate(), timeout=30)

            if process.returncode == 0 and output_path.exists():
                return output_path

            return None

        except Exception as e:
            logger.error(f"XLS conversion failed: {e}")
            return None
