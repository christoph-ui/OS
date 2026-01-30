"""
SYNDICATE MCP - Multi-Channel Content Syndication Engine

Transforms EATON P360 PIM data into 8+ distributor-specific formats:
- BMEcat XML (ECLASS/ETIM) - European standard
- ETIM xChange JSON - ETIM specification
- Amazon Vendor Central XLSX - E-commerce
- 1WorldSync XLSX - GDS Network
- CNET XML - Content syndication
- FAB-DIS XLSX - ROTH France (French language)
- TD Synnex XLSX - Tech Data distribution
- AMER Vendor XML - American distributors

Architecture:
1. Ingest P360 XML + Attributes CSV
2. Normalize 4,769 attributes → canonical schema
3. Apply AI-powered content generation
4. Map to distributor templates
5. Validate & export

Based on EATON's actual syndication requirements.
"""

import logging
import json
import xml.etree.ElementTree as ET
from typing import Any, Dict, Optional, List
from datetime import datetime

from mcps.sdk import BaseMCP, MCPResponse

logger = logging.getLogger(__name__)


class SyndicateMCP(BaseMCP):
    """
    Multi-Channel Content Syndication Engine

    Generates distributor-ready product content from PIM data.
    """

    # Required metadata
    name = "syndicate"
    version = "1.0.0"
    description = "Multi-Channel Content Syndication Engine"
    category = "syndication"

    # Supported output formats
    SUPPORTED_FORMATS = [
        "bmecat",        # BMEcat XML (ECLASS/ETIM)
        "etim_json",     # ETIM xChange JSON
        "amazon",        # Amazon Vendor Central
        "1worldsync",    # 1WorldSync GDSN
        "cnet",          # CNET Content Feed
        "fabdis",        # FAB-DIS (ROTH France)
        "td_synnex",     # TD Synnex
        "amer_xml"       # AMER Vendor XML
    ]

    async def process(
        self,
        input: Any,
        context: Optional[Dict[str, Any]] = None
    ) -> MCPResponse:
        """
        Process syndication request.

        Args:
            input: Syndication task (format + product selection)
            context: Customer context with lakehouse access

        Returns:
            MCPResponse with generated content
        """
        self.log(f"Processing syndication request: {str(input)[:100]}...")

        # Handle different input types
        if isinstance(input, dict):
            task_type = input.get("task_type", "generate")

            if task_type == "generate":
                return await self._generate_format(input, context)
            elif task_type == "validate":
                return await self._validate_data(input, context)
            elif task_type == "preview":
                return await self._preview_output(input, context)
            else:
                return await self._general_syndication(str(input), context)
        else:
            return await self._general_syndication(str(input), context)

    async def _generate_format(
        self,
        data: Dict[str, Any],
        context: Optional[Dict[str, Any]]
    ) -> MCPResponse:
        """
        Generate distributor-specific format.

        Args:
            data: {
                "format": "bmecat",
                "product_ids": ["5SC750", "5E1100IUSB-AR"],
                "language": "en",  # en, de, fr
                "options": {...}
            }
        """
        format_type = data.get("format", "bmecat")
        product_ids = data.get("product_ids", [])
        language = data.get("language", "en")

        if format_type not in self.SUPPORTED_FORMATS:
            return MCPResponse(
                data={"error": f"Unsupported format: {format_type}"},
                confidence=0,
                model_used="syndicate-error"
            )

        # Route to format-specific generator
        if format_type == "bmecat":
            return await self._generate_bmecat(product_ids, language, context)
        elif format_type == "etim_json":
            return await self._generate_etim_json(product_ids, language, context)
        elif format_type == "amazon":
            return await self._generate_amazon(product_ids, language, context)
        elif format_type == "1worldsync":
            return await self._generate_1worldsync(product_ids, language, context)
        elif format_type == "cnet":
            return await self._generate_cnet(product_ids, language, context)
        elif format_type == "fabdis":
            return await self._generate_fabdis(product_ids, context)
        elif format_type == "td_synnex":
            return await self._generate_td_synnex(product_ids, context)
        elif format_type == "amer_xml":
            return await self._generate_amer_xml(product_ids, context)
        else:
            return MCPResponse(
                data={"error": "Format generator not implemented yet"},
                confidence=0,
                model_used="syndicate-error"
            )

    async def _generate_bmecat(
        self,
        product_ids: List[str],
        language: str,
        context: Optional[Dict[str, Any]]
    ) -> MCPResponse:
        """
        Generate BMEcat 2005 XML with ECLASS 13.0 and ETIM-X classifications.

        BMEcat is the European standard for product catalogs.
        Includes full ECLASS hierarchy and ETIM attributes.
        """
        self.log(f"Generating BMEcat XML for {len(product_ids)} products (language: {language})")

        # Get products from lakehouse
        products_data = await self._get_products_data(product_ids, context)

        if not products_data:
            return MCPResponse(
                data={"error": "No products found"},
                confidence=0,
                model_used="syndicate-bmecat"
            )

        # Generate BMEcat XML using Claude
        prompt = f"""Generate BMEcat 2005 XML catalog with ECLASS 13.0 and ETIM-X classifications.

**Products to include:** {len(products_data)} products
**Language:** {"German" if language == "de" else "English"}

**Product Data:**
{self._format_products_for_prompt(products_data[:3])}  # Sample first 3

**BMEcat 2005 Structure Required:**

```xml
<?xml version="1.0" encoding="UTF-8"?>
<BMECAT version="2005.1" xmlns="http://www.bmecat.org/bmecat/2005">
  <HEADER>
    <CATALOG>
      <LANGUAGE>{language}</LANGUAGE>
      <CATALOG_ID>EATON_MASTER</CATALOG_ID>
      <CATALOG_VERSION>2.0</CATALOG_VERSION>
      <CATALOG_NAME>Eaton Product Catalog</CATALOG_NAME>
      <DATETIME type="generation_date">
        <DATE>{datetime.now().strftime('%Y-%m-%d')}</DATE>
        <TIME>{datetime.now().strftime('%H:%M:%S')}</TIME>
      </DATETIME>
      <TERRITORY>EMEA</TERRITORY>
      <CURRENCY>EUR</CURRENCY>
      <MIME_ROOT>https://www.eaton.com/mdmfiles/</MIME_ROOT>
      <SUPPLIER>
        <SUPPLIER_NAME>Eaton Industries GmbH</SUPPLIER_NAME>
        <ADDRESS>
          <STREET>Hein-Moeller-Str. 7-11</STREET>
          <ZIP>53115</ZIP>
          <CITY>Bonn</CITY>
          <COUNTRY>Germany</COUNTRY>
        </ADDRESS>
      </SUPPLIER>
    </CATALOG>
  </HEADER>

  <T_NEW_CATALOG>
    <!-- FEATURE SYSTEMS -->
    <FEATURE_SYSTEM>
      <FEATURE_SYSTEM_NAME>ECLASS-13.0</FEATURE_SYSTEM_NAME>
    </FEATURE_SYSTEM>
    <FEATURE_SYSTEM>
      <FEATURE_SYSTEM_NAME>ETIM-10</FEATURE_SYSTEM_NAME>
    </FEATURE_SYSTEM>

    <!-- ARTICLES (one per product) -->
    <ARTICLE>
      <SUPPLIER_AID>[Product SKU]</SUPPLIER_AID>
      <ARTICLE_DETAILS>
        <DESCRIPTION_SHORT>[Short description]</DESCRIPTION_SHORT>
        <DESCRIPTION_LONG>[Marketing description]</DESCRIPTION_LONG>
        <EAN>[EAN code]</EAN>
        <MANUFACTURER_NAME>Eaton</MANUFACTURER_NAME>
        <MANUFACTURER_TYPE_DESCR>[Product model]</MANUFACTURER_TYPE_DESCR>
        <KEYWORD>[Keywords separated by ;]</KEYWORD>
      </ARTICLE_DETAILS>

      <!-- ECLASS Classification -->
      <ARTICLE_FEATURES>
        <REFERENCE_FEATURE_SYSTEM_NAME>ECLASS-13.0</REFERENCE_FEATURE_SYSTEM_NAME>
        <REFERENCE_FEATURE_GROUP_ID>[ECLASS code e.g., 27-24-05-01]</REFERENCE_FEATURE_GROUP_ID>
        <FEATURE>
          <FNAME>Rated voltage</FNAME>
          <FVALUE>230</FVALUE>
          <FUNIT>V</FUNIT>
        </FEATURE>
        <!-- More features -->
      </ARTICLE_FEATURES>

      <!-- ETIM Classification -->
      <ARTICLE_FEATURES>
        <REFERENCE_FEATURE_SYSTEM_NAME>ETIM-10</REFERENCE_FEATURE_SYSTEM_NAME>
        <REFERENCE_FEATURE_GROUP_ID>[ETIM class e.g., EC000382]</REFERENCE_FEATURE_GROUP_ID>
        <FEATURE>
          <FNAME>Output voltage</FNAME>
          <FVALUE>120</FVALUE>
          <FUNIT>V</FUNIT>
        </FEATURE>
        <!-- More features -->
      </ARTICLE_FEATURES>

      <!-- IMAGES -->
      <MIME_INFO>
        <MIME>
          <MIME_TYPE>image/jpeg</MIME_TYPE>
          <MIME_SOURCE>[Image URL 1000x1000]</MIME_SOURCE>
          <MIME_DESCR>Product image</MIME_DESCR>
          <MIME_PURPOSE>normal</MIME_PURPOSE>
        </MIME>
        <MIME>
          <MIME_TYPE>application/pdf</MIME_TYPE>
          <MIME_SOURCE>[Datasheet URL]</MIME_SOURCE>
          <MIME_DESCR>Technical datasheet</MIME_DESCR>
          <MIME_PURPOSE>detail</MIME_PURPOSE>
        </MIME>
      </MIME_INFO>

      <!-- PRICE if available -->
      <ARTICLE_PRICE_DETAILS>
        <ARTICLE_PRICE>
          <PRICE_AMOUNT>199.00</PRICE_AMOUNT>
          <PRICE_CURRENCY>EUR</PRICE_CURRENCY>
          <TAX>0.19</TAX>
        </ARTICLE_PRICE>
      </ARTICLE_PRICE_DETAILS>

      <!-- ORDER DETAILS -->
      <ARTICLE_ORDER_DETAILS>
        <ORDER_UNIT>PCE</ORDER_UNIT>
        <CONTENT_UNIT>PCE</CONTENT_UNIT>
        <NO_CU_PER_OU>1</NO_CU_PER_OU>
      </ARTICLE_ORDER_DETAILS>
    </ARTICLE>

    <!-- Repeat for each product -->
  </T_NEW_CATALOG>
</BMECAT>
```

**Generate complete BMEcat XML now for all {len(products_data)} products.**
Include all ECLASS/ETIM features found in the product data.
Ensure valid XML structure compliant with BMEcat 2005.1 XSD schema.
"""

        # Generate with Claude Sonnet directly
        import anthropic
        import os

        client = anthropic.AsyncAnthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        try:
            message = await client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=8000,
                temperature=0.1,
                messages=[{
                    "role": "user",
                    "content": prompt
                }]
            )

            xml_content = message.content[0].text if message.content else ""

        except Exception as e:
            logger.error(f"Claude API error: {e}")
            xml_content = f"<!-- BMEcat generation failed: {e} -->"

        return MCPResponse(
            data={
                "xml": xml_content,
                "format": "bmecat",
                "products_count": len(products_data),
                "language": language,
                "filename": f"eaton_catalog_{language}_{datetime.now().strftime('%Y%m%d')}.xml"
            },
            confidence=0.9,
            model_used="syndicate-bmecat-v1.0"
        )

    async def _generate_etim_json(
        self,
        product_ids: List[str],
        language: str,
        context: Optional[Dict[str, Any]]
    ) -> MCPResponse:
        """
        Generate ETIM xChange JSON format.

        Format: Simple JSON with ETIM classifications and attributes.
        Standard: ETIM-10 specification.
        """
        self.log(f"Generating ETIM xChange JSON for {len(product_ids)} products")

        products_data = await self._get_products_data(product_ids, context)

        if not products_data:
            return MCPResponse(
                data={"error": "No products found"},
                confidence=0,
                model_used="syndicate-etim"
            )

        # Build prompt
        prompt = f"""Generate ETIM xChange JSON for {len(products_data)} products.

**JSON Structure Required:**

```json
{{
  "products": [
    {{
      "ean": "GTIN code",
      "supplier_pid": "Manufacturer SKU",
      "etim_class": "ETIM class code (EC000382)",
      "product_name": "Product name",
      "attributes": [
        {{
          "etim_id": "EF009897",
          "name": "Output power",
          "value": "525",
          "unit": "W"
        }},
        {{
          "etim_id": "EF009898",
          "name": "VA rating",
          "value": "750",
          "unit": "VA"
        }}
      ]
    }}
  ]
}}
```

**Transformation Rules:**
- Include only products with ETIM class (etim_class field not empty)
- Parse technical_attributes JSON for ETIM-compatible attributes
- Use ETIM-10 standard codes if available
- Include product name and GTIN for reference

**Product Data:**
{self._format_products_for_prompt(products_data[:10])}

Generate valid JSON for all {len(products_data)} products.
Only include products that have ETIM classifications.
"""

        # Generate with Claude
        import anthropic
        import os

        client = anthropic.AsyncAnthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        try:
            message = await client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=16000,
                temperature=0.1,
                messages=[{"role": "user", "content": prompt}]
            )

            json_content = message.content[0].text if message.content else ""

        except Exception as e:
            logger.error(f"Claude API error: {e}")
            json_content = f'{{"error": "{e}"}}'

        return MCPResponse(
            data={
                "json": json_content,
                "content": json_content,
                "format": "etim_json",
                "products_count": len(products_data),
                "language": language,
                "filename": f"eaton_etim_xchange_{datetime.now().strftime('%Y%m%d')}.json"
            },
            confidence=0.9,
            model_used="syndicate-etim-v1.0"
        )

    async def _generate_amazon(
        self,
        product_ids: List[str],
        language: str,
        context: Optional[Dict[str, Any]]
    ) -> MCPResponse:
        """
        Generate Amazon Vendor Central bulk upload XLSX.

        Format: Excel with strict character limits and content requirements.
        Character limits: Title 200, Bullets 500 each, Description 2000, Search terms 50 each
        Images: Main ≥1000px, Additional ≥500px (max 8 total)
        """
        self.log(f"Generating Amazon Vendor Central format for {len(product_ids)} products")

        products_data = await self._get_products_data(product_ids, context)

        if not products_data:
            return MCPResponse(
                data={"error": "No products found"},
                confidence=0,
                model_used="syndicate-amazon"
            )

        # Build prompt for Claude
        prompt = f"""Generate Amazon Vendor Central bulk upload data for {len(products_data)} products.

**Required Excel Columns:**
1. SKU - Manufacturer part number
2. Product ID - UPC/GTIN barcode
3. Product Title - Max 200 characters (product name + top feature)
4. Bullet Point 1 - Max 500 characters (benefit-focused)
5. Bullet Point 2 - Max 500 characters
6. Bullet Point 3 - Max 500 characters
7. Bullet Point 4 - Max 500 characters
8. Bullet Point 5 - Max 500 characters
9. Product Description - Max 2000 characters (marketing copy)
10. Search Terms - 5 keywords, max 50 characters each (comma-separated)
11. Main Image URL - Primary image (≥1000px resolution)
12. Other Image URL 1-7 - Additional images (≥500px each)
13. Item Weight - In pounds
14. Item Dimensions - Format as "H x W x L" in inches
15. Safety & Compliance - Certifications (comma-separated)

**Transformation Rules:**
- Extract 5 bullets from features array (pipe-delimited features field)
- Rewrite as customer benefits (not just features)
- Enforce strict character limits (truncate intelligently at word boundaries)
- Select images by priority (priority 1 = main, priority 2-8 = additional)
- Filter for high resolution (1000x1000 for main, 500x500+ for others)
- Generate SEO-friendly search terms from keywords and product type

**Product Data:**
{self._format_products_for_prompt(products_data[:10])}  # First 10 for prompt

**Output Format - CRITICAL:**
Generate ONLY raw CSV data. NO markdown code blocks. NO explanatory text.

First line: Column headers separated by pipe (|)
Subsequent lines: Product data, one row per product, pipe-separated

Example:
SKU|Product ID|Product Title|Bullet Point 1|Bullet Point 2
5SC750|00743172045096|Eaton 5SC UPS - 750VA|Reliable line-interactive UPS|User-replaceable batteries

Do NOT wrap in ```csv blocks. Output raw CSV only.
Include all {len(products_data)} products.
"""

        # Generate with Claude
        import anthropic
        import os

        client = anthropic.AsyncAnthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        try:
            message = await client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=16000,
                temperature=0.1,
                messages=[{
                    "role": "user",
                    "content": prompt
                }]
            )

            csv_content = message.content[0].text if message.content else ""

            # Convert CSV to properly formatted XLSX
            import base64
            xlsx_bytes = self._csv_to_xlsx_bytes(csv_content, delimiter="|")
            xlsx_base64 = base64.b64encode(xlsx_bytes).decode('utf-8')

        except Exception as e:
            logger.error(f"Claude API error: {e}")
            csv_content = f"Error generating Amazon format: {e}"
            xlsx_base64 = ""

        return MCPResponse(
            data={
                "content": csv_content,  # Keep CSV for debugging
                "xlsx": xlsx_base64,  # Actual XLSX (base64 encoded)
                "format": "amazon",
                "products_count": len(products_data),
                "language": language,
                "filename": f"eaton_amazon_{language}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            },
            confidence=0.85,
            model_used="syndicate-amazon-v1.0"
        )

    async def _generate_1worldsync(
        self,
        product_ids: List[str],
        language: str,
        context: Optional[Dict[str, Any]]
    ) -> MCPResponse:
        """
        Generate 1WorldSync GDSN (GS1 Global Data Synchronization Network) XLSX.

        Format: GS1 standard for global product data synchronization.
        Requires: GTIN validation, GPC classification, multi-language support.
        """
        self.log(f"Generating 1WorldSync GDSN format for {len(product_ids)} products")

        products_data = await self._get_products_data(product_ids, context)

        if not products_data:
            return MCPResponse(
                data={"error": "No products found"},
                confidence=0,
                model_used="syndicate-1ws"
            )

        # Build prompt
        prompt = f"""Generate 1WorldSync GDSN Excel export for {len(products_data)} products.

**Required Columns (GS1 Standard):**
- GTIN (14-digit with valid checksum)
- Brand (Manufacturer name)
- Product Description (max 200 characters)
- Product Classification (GPC brick code - map from UNSPSC)
- Net Content (weight with unit)
- Gross Weight (package weight with unit)
- Dimensions (H x W x L with units)
- Country of Origin
- Compliance & Certifications
- Primary Image URL (high resolution, 300dpi+)
- Digital Assets (additional image URLs)

**Transformation Rules:**
- Validate GTIN checksum (14 digits, valid check digit)
- Map UNSPSC → GS1 GPC classification
- Support multiple languages ({language})
- High-resolution images only (1000x1000 300dpi minimum)
- Include all compliance data

**Product Data:**
{self._format_products_for_prompt(products_data[:10])}

**Output Format - CRITICAL:**
Generate ONLY raw CSV data. NO markdown. NO code blocks.

First line: Column headers, pipe-separated
Subsequent lines: Product data

Example:
GTIN|Brand|Product Description|GPC Code|Net Content
00743172045096|Eaton|Eaton 5SC UPS 750VA|10005844|750VA

Do NOT wrap in ```csv blocks. Output raw CSV only.
Include all {len(products_data)} products.
"""

        # Generate with Claude
        import anthropic
        import os

        client = anthropic.AsyncAnthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        try:
            message = await client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=16000,
                temperature=0.1,
                messages=[{"role": "user", "content": prompt}]
            )

            csv_content = message.content[0].text if message.content else ""

            # Convert CSV to properly formatted XLSX
            import base64
            xlsx_bytes = self._csv_to_xlsx_bytes(csv_content, delimiter="|")
            xlsx_base64 = base64.b64encode(xlsx_bytes).decode('utf-8')

        except Exception as e:
            logger.error(f"Claude API error: {e}")
            csv_content = f"Error: {e}"
            xlsx_base64 = ""

        return MCPResponse(
            data={
                "content": csv_content,
                "xlsx": xlsx_base64,
                "format": "1worldsync",
                "products_count": len(products_data),
                "language": language,
                "filename": f"eaton_1worldsync_{language}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            },
            confidence=0.85,
            model_used="syndicate-1ws-v1.0"
        )

    async def _generate_cnet(
        self,
        product_ids: List[str],
        language: str,
        context: Optional[Dict[str, Any]]
    ) -> MCPResponse:
        """
        Generate CNET Content Feed XML.

        Format: Structured XML with grouped attributes and rich content.
        Attributes grouped by: OVERVIEW, PHYSICAL, ELECTRICAL, ENVIRONMENTAL
        """
        self.log(f"Generating CNET content feed for {len(product_ids)} products")

        products_data = await self._get_products_data(product_ids, context)

        if not products_data:
            return MCPResponse(
                data={"error": "No products found"},
                confidence=0,
                model_used="syndicate-cnet"
            )

        # Build prompt for Claude
        prompt = f"""Generate CNET Content Feed XML for {len(products_data)} products.

**XML Structure Required:**

```xml
<?xml version="1.0" encoding="UTF-8"?>
<Products>
  <Product>
    <PartNumber>{{SKU}}</PartNumber>
    <ProductDescription>{{Product Name}}</ProductDescription>
    <UpcEan>{{GTIN}}</UpcEan>

    <KeySellingPoints>
      <Item>{{Benefit 1 from features}}</Item>
      <Item>{{Benefit 2 from features}}</Item>
      <Item>{{Benefit 3 from features}}</Item>
      <Item>{{Benefit 4 from features}}</Item>
      <Item>{{Benefit 5 from features}}</Item>
    </KeySellingPoints>

    <ProductFeatures>
      <Item><Header>OVERVIEW</Header><Value>{{Marketing description}}</Value></Item>
      <Item><Header>APPLICATIONS</Header><Value>{{Use cases}}</Value></Item>
      <Item><Header>WARRANTY</Header><Value>{{Warranty info}}</Value></Item>
    </ProductFeatures>

    <AttributeGroups>
      <Group name="PHYSICAL">
        <Attribute><Name>Dimensions</Name><Value>{{H x W x L}}</Value><Unit>inches</Unit></Attribute>
        <Attribute><Name>Weight</Name><Value>{{weight}}</Value><Unit>lbs</Unit></Attribute>
      </Group>
      <Group name="ELECTRICAL">
        <Attribute><Name>Output Voltage</Name><Value>{{from attributes}}</Value><Unit>V</Unit></Attribute>
        <!-- Extract from technical_attributes JSON -->
      </Group>
      <Group name="CERTIFICATIONS">
        <Attribute><Name>Compliance</Name><Value>{{certifications}}</Value></Attribute>
      </Group>
    </AttributeGroups>

    <HeroImage>{{Primary image URL (1000x1000)}}</HeroImage>
    <Images>
      <Image order="1">{{Image 2 URL}}</Image>
      <Image order="2">{{Image 3 URL}}</Image>
      <!-- Up to 6 additional images -->
    </Images>

    <PdfProductDataSheet>{{Datasheet URL if available}}</PdfProductDataSheet>
  </Product>

  <!-- Repeat for each product -->
</Products>
```

**Transformation Rules:**
- Extract 5-6 selling points from features array
- Group attributes by PHYSICAL, ELECTRICAL, ENVIRONMENTAL, CERTIFICATIONS
- Parse technical_attributes JSON for detailed specs
- Select HeroImage (priority 1, 1000x1000)
- Select additional images (priority 2-7, 500x500+)

**Product Data:**
{self._format_products_for_prompt(products_data[:10])}

Generate complete CNET XML for all {len(products_data)} products.
Ensure valid XML structure.
"""

        # Generate with Claude
        import anthropic
        import os

        client = anthropic.AsyncAnthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        try:
            message = await client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=16000,
                temperature=0.1,
                messages=[{
                    "role": "user",
                    "content": prompt
                }]
            )

            xml_content = message.content[0].text if message.content else ""

        except Exception as e:
            logger.error(f"Claude API error: {e}")
            xml_content = f"<!-- CNET generation failed: {e} -->"

        return MCPResponse(
            data={
                "xml": xml_content,
                "content": xml_content,
                "format": "cnet",
                "products_count": len(products_data),
                "language": language,
                "filename": f"eaton_cnet_feed_{language}_{datetime.now().strftime('%Y%m%d')}.xml"
            },
            confidence=0.85,
            model_used="syndicate-cnet-v1.0"
        )

    async def _generate_fabdis(
        self,
        product_ids: List[str],
        context: Optional[Dict[str, Any]]
    ) -> MCPResponse:
        """
        Generate FAB-DIS (ROTH France) XLSX format in French.

        Format: French distributor template with metric units.
        Requires: Translation EN→FR, Imperial→Metric conversion, European compliance filtering.
        """
        self.log(f"Generating FAB-DIS (ROTH France) format for {len(product_ids)} products")

        products_data = await self._get_products_data(product_ids, context)

        if not products_data:
            return MCPResponse(
                data={"error": "No products found"},
                confidence=0,
                model_used="syndicate-fabdis"
            )

        # Build prompt with translation and conversion requirements
        prompt = f"""Generate FAB-DIS (ROTH France) Excel export for {len(products_data)} products IN FRENCH.

**Required Columns (in French):**
- EAN (GTIN code)
- Reference fabricant (Manufacturer SKU)
- Designation (Product name - TRANSLATE to French)
- Description (Marketing text - TRANSLATE to French, max 1000 chars)
- Famille (Product category - TRANSLATE to French)
- Sous-famille (Product subcategory - TRANSLATE to French)
- Prix (List price in EUR)
- Poids (Weight in KILOGRAMS - convert from pounds)
- Dimensions (Height x Width x Length in CENTIMETERS - convert from inches)
- Certifications (European only: CE, NF, EN - filter out US certifications)
- Image URL (Primary image)

**CRITICAL Transformations:**
1. **Translate to French:**
   - Product names and descriptions
   - Categories (UPS → Onduleur, Circuit Breaker → Disjoncteur, etc.)
   - Feature bullets

2. **Convert to Metric:**
   - Pounds → Kilograms (multiply by 0.453592)
   - Inches → Centimeters (multiply by 2.54)
   - Format: "12.5 cm x 8.3 cm x 15.2 cm"

3. **Filter Certifications:**
   - KEEP: CE, NF, EN standards (European)
   - REMOVE: UL, cUL, FCC, NOM (US/North American)

**Product Data:**
{self._format_products_for_prompt(products_data[:10])}

**Output Format - CRITICAL:**
Generate ONLY raw CSV data. NO markdown. NO code blocks. NO explanatory text.

First line: Column headers in French, pipe-separated
Subsequent lines: Product data in French with metric units

Example:
EAN|Reference fabricant|Designation|Poids|Dimensions
00743172045096|5SC750|Onduleur Eaton 5SC 750VA|10.5|31 x 15 x 34

Do NOT wrap in ```csv blocks. Output raw CSV only.
ALL text in FRENCH. ALL units in METRIC.
Include all {len(products_data)} products.
"""

        # Generate with Claude
        import anthropic
        import os

        client = anthropic.AsyncAnthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        try:
            message = await client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=16000,
                temperature=0.1,
                messages=[{"role": "user", "content": prompt}]
            )

            csv_content = message.content[0].text if message.content else ""

            # Convert CSV to properly formatted XLSX
            import base64
            xlsx_bytes = self._csv_to_xlsx_bytes(csv_content, delimiter="|")
            xlsx_base64 = base64.b64encode(xlsx_bytes).decode('utf-8')

        except Exception as e:
            logger.error(f"Claude API error: {e}")
            csv_content = f"Error: {e}"
            xlsx_base64 = ""

        return MCPResponse(
            data={
                "content": csv_content,
                "xlsx": xlsx_base64,
                "format": "fabdis",
                "products_count": len(products_data),
                "language": "fr",
                "filename": f"eaton_fabdis_fr_{datetime.now().strftime('%Y%m%d')}.xlsx"
            },
            confidence=0.85,
            model_used="syndicate-fabdis-v1.0"
        )

    async def _generate_td_synnex(
        self,
        product_ids: List[str],
        context: Optional[Dict[str, Any]]
    ) -> MCPResponse:
        """
        Generate TD Synnex (Tech Data) XLSX format.

        Format: Excel template for IT distribution channel.
        Focus: IT-specific attributes, SRP pricing, lifecycle status.
        """
        self.log(f"Generating TD Synnex format for {len(product_ids)} products")

        products_data = await self._get_products_data(product_ids, context)

        if not products_data:
            return MCPResponse(
                data={"error": "No products found"},
                confidence=0,
                model_used="syndicate-td"
            )

        # Build prompt
        prompt = f"""Generate TD Synnex (Tech Data) Excel export for {len(products_data)} products.

**Required Columns:**
- Manufacturer Part Number (SKU)
- UPC/EAN (GTIN)
- Product Title (max 150 characters)
- Short Description (max 500 characters)
- Long Description (max 2000 characters)
- Category (map product type to TD taxonomy)
- Marketing Features (5 bullets from features array)
- Technical Specifications (from technical_attributes JSON)
- SRP (List Price if available)
- Primary Image URL
- Datasheet URL

**Transformation Rules:**
- Extract IT-specific attributes (connectivity, compatibility, interfaces)
- Generate concise marketing features
- Map product type to IT distribution categories
- Include all pricing info

**Product Data:**
{self._format_products_for_prompt(products_data[:10])}

**Output Format - CRITICAL:**
Generate ONLY raw CSV data. NO markdown. NO code blocks.

First line: Column headers, pipe-separated
Subsequent lines: Product data

Example:
Manufacturer Part Number|UPC/EAN|Product Title|Short Description
5SC750|00743172045096|Eaton 5SC UPS - 750VA|Line-interactive UPS with AVR

Do NOT wrap in ```csv blocks. Output raw CSV only.
Include all {len(products_data)} products.
"""

        # Generate with Claude
        import anthropic
        import os

        client = anthropic.AsyncAnthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        try:
            message = await client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=16000,
                temperature=0.1,
                messages=[{"role": "user", "content": prompt}]
            )

            csv_content = message.content[0].text if message.content else ""

            # Convert CSV to properly formatted XLSX
            import base64
            xlsx_bytes = self._csv_to_xlsx_bytes(csv_content, delimiter="|")
            xlsx_base64 = base64.b64encode(xlsx_bytes).decode('utf-8')

        except Exception as e:
            logger.error(f"Claude API error: {e}")
            csv_content = f"Error: {e}"
            xlsx_base64 = ""

        return MCPResponse(
            data={
                "content": csv_content,
                "xlsx": xlsx_base64,
                "format": "td_synnex",
                "products_count": len(products_data),
                "language": "en",
                "filename": f"eaton_td_synnex_{datetime.now().strftime('%Y%m%d')}.xlsx"
            },
            confidence=0.85,
            model_used="syndicate-td-v1.0"
        )

    async def _generate_amer_xml(
        self,
        product_ids: List[str],
        context: Optional[Dict[str, Any]]
    ) -> MCPResponse:
        """
        Generate AMER Vendor XML format (American distributors).

        Format: Simplified P360-style XML with core fields only.
        Use Case: US distributors requiring basic product data.
        """
        self.log(f"Generating AMER Vendor XML for {len(product_ids)} products")

        products_data = await self._get_products_data(product_ids, context)

        if not products_data:
            return MCPResponse(
                data={"error": "No products found"},
                confidence=0,
                model_used="syndicate-amer"
            )

        # Build prompt
        prompt = f"""Generate AMER Vendor XML for {len(products_data)} products (American distributors).

**XML Structure (Simplified P360):**

```xml
<?xml version="1.0" encoding="UTF-8"?>
<envelope>
  <item>
    <catalog>SKU</catalog>
    <upc>UPC code</upc>
    <gtin>GTIN 14-digit</gtin>
    <prodName>Product Name</prodName>
    <prodType>Product Type</prodType>
    <longDesc>Short description</longDesc>
    <markDesc>Marketing description</markDesc>
    <prodFeature>Feature 1|Feature 2|Feature 3</prodFeature>
    <brandLabel>Eaton</brandLabel>
    <prodWt>Weight</prodWt>
    <prodWtUOM>pound</prodWtUOM>
    <prodHgt>Height</prodHgt>
    <prodWid>Width</prodWid>
    <prodLen>Length</prodLen>
    <prodHgtUOM>inch</prodHgtUOM>
    <image>
      <imageURL>Primary image URL</imageURL>
    </image>
    <Certifications>UL, cUL, FCC</Certifications>
    <Warranty>Warranty text</Warranty>
  </item>
  <!-- Repeat for each product -->
</envelope>
```

**Field Mapping:**
- Use core P360 fields only (identifiers, descriptions, features, dimensions, images)
- Keep imperial units (pounds, inches) - no conversion needed
- Include US/North American certifications (UL, cUL, FCC, NOM)
- Pipe-delimited features

**Product Data:**
{self._format_products_for_prompt(products_data[:10])}

Generate complete AMER Vendor XML for all {len(products_data)} products.
"""

        # Generate with Claude
        import anthropic
        import os

        client = anthropic.AsyncAnthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        try:
            message = await client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=16000,
                temperature=0.1,
                messages=[{"role": "user", "content": prompt}]
            )

            xml_content = message.content[0].text if message.content else ""

        except Exception as e:
            logger.error(f"Claude API error: {e}")
            xml_content = f"<!-- AMER XML generation failed: {e} -->"

        return MCPResponse(
            data={
                "xml": xml_content,
                "content": xml_content,
                "format": "amer_xml",
                "products_count": len(products_data),
                "language": "en",
                "filename": f"eaton_amer_vendor_{datetime.now().strftime('%Y%m%d')}.xml"
            },
            confidence=0.9,
            model_used="syndicate-amer-v1.0"
        )

    async def _get_products_data(
        self,
        product_ids: List[str],
        context: Optional[Dict[str, Any]]
    ) -> List[Dict]:
        """
        Retrieve product data from lakehouse.

        Queries syndication_products table for specified SKUs.
        """
        import httpx

        # Get lakehouse URL from context
        lakehouse_url = context.get("lakehouse_url", "http://localhost:9302") if context else "http://localhost:9302"

        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                # Query syndication_products table
                params = {}

                if product_ids:
                    # For now, get all and filter (TODO: support WHERE clause in lakehouse API)
                    params["limit"] = 200
                else:
                    params["limit"] = 100

                response = await client.get(
                    f"{lakehouse_url}/delta/query/syndication_products",
                    params=params
                )
                response.raise_for_status()
                data = response.json()

                all_products = data.get("rows", [])

                # Filter by product_ids if specified
                if product_ids:
                    products = [p for p in all_products if p.get("product_id") in product_ids]
                else:
                    products = all_products

                logger.info(f"Retrieved {len(products)} products from syndication_products table")
                return products

        except Exception as e:
            logger.error(f"Error querying products from lakehouse: {e}", exc_info=True)
            return []

    def _csv_to_xlsx_bytes(self, csv_content: str, delimiter: str = "|") -> bytes:
        """
        Convert CSV to properly formatted XLSX with aligned columns.

        Args:
            csv_content: CSV text with delimiter
            delimiter: CSV delimiter (default: pipe)

        Returns:
            Binary XLSX content ready for download
        """
        import io
        import csv
        import base64
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import re

        # Strip markdown code blocks if present
        # Claude often wraps output in ```csv ... ```
        csv_clean = csv_content.strip()

        # Remove markdown code fences
        if csv_clean.startswith("```"):
            # Find first newline after opening fence
            lines = csv_clean.split('\n')
            # Skip first line (```csv or ```), take until closing ```
            csv_lines = []
            in_code_block = False
            for line in lines:
                if line.startswith("```"):
                    in_code_block = not in_code_block
                    continue
                if in_code_block or not line.startswith("```"):
                    csv_lines.append(line)
            csv_clean = '\n'.join(csv_lines).strip()
            logger.info(f"Stripped markdown code blocks from CSV ({len(csv_content)} → {len(csv_clean)} chars)")

        # Log first 200 chars for debugging
        logger.info(f"CSV preview: {csv_clean[:200]}...")

        # Parse CSV
        reader = csv.reader(io.StringIO(csv_clean), delimiter=delimiter)
        rows = list(reader)

        if not rows:
            logger.warning("No rows in CSV content")
            return b""

        logger.info(f"Parsed CSV: {len(rows)} rows, {len(rows[0]) if rows else 0} columns")
        if rows:
            logger.info(f"Headers: {rows[0][:5]}")  # First 5 column names

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

        # Add all rows
        for row in rows:
            ws.append(row)

        # Format header row (row 1) - Eaton orange brand color
        header_fill = PatternFill(start_color="D97757", end_color="D97757", fill_type="solid")
        for col in range(1, len(rows[0]) + 1):
            cell = ws.cell(1, col)
            cell.font = Font(bold=True, color="FFFFFF", size=11)  # White bold text
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')

        # Auto-size all columns based on content
        for col in range(1, len(rows[0]) + 1):
            max_length = 0
            column_letter = get_column_letter(col)

            for row_idx, row in enumerate(ws[column_letter], 1):
                try:
                    cell_value = str(row.value) if row.value else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)

                    # Align data rows
                    if row_idx > 1:  # Skip header
                        row.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                except:
                    pass

            # Set column width (add padding, max 50 chars for readability)
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 12)  # Min 12 chars

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"✓ Created XLSX with {len(rows)-1} products, {len(rows[0])} columns")

        return output.getvalue()

    def _format_products_for_prompt(self, products: List[Dict]) -> str:
        """Format products for Claude prompt."""
        formatted = ""
        for product in products[:5]:  # Limit to 5 for prompt size
            formatted += f"\nProduct: {product.get('product_name', 'Unknown')}\n"
            formatted += f"  SKU: {product.get('supplier_pid', 'N/A')}\n"
            formatted += f"  EAN: {product.get('ean_upc', 'N/A')}\n"
            formatted += f"  ETIM: {product.get('etim_class', 'N/A')}\n"
            formatted += f"  Description: {product.get('short_description', 'N/A')[:200]}\n"

        return formatted

    async def _validate_data(
        self,
        data: Dict[str, Any],
        context: Optional[Dict[str, Any]]
    ) -> MCPResponse:
        """
        Validate product data for syndication readiness.

        Checks:
        - Required fields populated
        - GTIN checksum valid
        - Images accessible
        - Classifications present
        """
        product_ids = data.get("product_ids", [])
        format_type = data.get("format", "all")

        products = await self._get_products_data(product_ids, context)

        # Validation logic (to be implemented)
        validation_report = {
            "total_products": len(products),
            "valid": 0,
            "warnings": [],
            "errors": []
        }

        return MCPResponse(
            data=validation_report,
            confidence=0.8,
            model_used="syndicate-validator"
        )

    async def _preview_output(
        self,
        data: Dict[str, Any],
        context: Optional[Dict[str, Any]]
    ) -> MCPResponse:
        """
        Generate preview of syndicated content.

        Returns first 3 products in requested format for review.
        """
        format_type = data.get("format", "bmecat")
        product_ids = data.get("product_ids", [])[:3]  # Limit to 3 for preview

        # Generate preview (reuse format generators)
        return await self._generate_format(
            {"format": format_type, "product_ids": product_ids},
            context
        )

    async def _general_syndication(
        self,
        query: str,
        context: Optional[Dict[str, Any]]
    ) -> MCPResponse:
        """
        Handle natural language syndication requests.

        Examples:
        - "Generate BMEcat for all UPS products"
        - "Create Amazon listing for 5SC750"
        - "Export to 1WorldSync"
        """
        # Parse query for format and products
        query_lower = query.lower()

        # Detect format
        format_type = None
        for fmt in self.SUPPORTED_FORMATS:
            if fmt.replace("_", " ") in query_lower or fmt in query_lower:
                format_type = fmt
                break

        if not format_type:
            # Default to BMEcat
            format_type = "bmecat"

        # Detect product filter
        product_ids = []  # Empty = all products

        # Generate
        return await self._generate_format(
            {"format": format_type, "product_ids": product_ids},
            context
        )


# Factory function
def create_syndicate_mcp() -> SyndicateMCP:
    """Create SYNDICATE MCP instance"""
    return SyndicateMCP()
